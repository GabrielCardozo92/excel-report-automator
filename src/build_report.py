import pandas as pd
from pathlib import Path
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
OUTPUT_DIR = Path(__file__).resolve().parent.parent / "output"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
CURRENCY_FMT = '"$"#,##0.00'

def load_clean_data():
    """Load the cleaned sales data (fails clearly if cleaning wasn't run)."""
    file_path = DATA_DIR / "cleaned_sales.xlsx"
    if not file_path.exists():
        raise FileNotFoundError(f"Run clean_data.py first — missing: {file_path}")
    df = pd.read_excel(file_path)
    df["Date"] = pd.to_datetime(df["Date"])
    return df

def aggregate_by(df, column):
    """Revenue/Units/Orders grouped by any column, best sellers first."""
    return (df.groupby(column)
              .agg(Revenue=("Total", "sum"),
                   Units=("Quantity", "sum"),
                   Orders=("Total", "count"))
              .sort_values("Revenue", ascending=False)
              .reset_index())

def monthly_trend(df):
    """Revenue and Orders per month, in chronological order."""
    df = df.copy()
    df["Month"] = df["Date"].dt.to_period("M")
    return (df.groupby("Month")
              .agg(Revenue=("Total", "sum"),
                   Orders=("Total", "count"))
              .sort_values("Month")
              .reset_index())

def compute_kpis(df):
    """Executive summary numbers as a dict {label: value}."""
    return {
        "Total Revenue": round(df["Total"].sum(), 2),
        "Total Units": df["Quantity"].sum(),
        "Total Orders": df["Total"].count(),
        "Average Order Value": round(df["Total"].mean(), 2),
        "Top Product": df.groupby("Product")["Total"].sum().idxmax(),
        "Top Region": df.groupby("Region")["Total"].sum().idxmax(),
        "Top Seller": df.groupby("Seller")["Total"].sum().idxmax(),
        "Period": f"{df['Date'].min().strftime('%Y-%m-%d')} to {df['Date'].max().strftime('%Y-%m-%d')}"
    }

def _format_sheet(ws, currency_cols=()):
    """Style header row, auto-fit column widths, apply currency format."""
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        width = max(len(str(c.value)) for c in ws[letter] if c.value is not None)
        ws.column_dimensions[letter].width = max(12, width + 2)
        if letter in currency_cols:
            for cell in ws[letter][1:]:
                cell.number_format = CURRENCY_FMT


def build_report(df, kpis, by_product, by_region, by_seller, by_month):
    """Write the formatted multi-sheet Excel report with an embedded chart."""
    OUTPUT_DIR.mkdir(exist_ok=True)
    path = OUTPUT_DIR / "sales_report.xlsx"
    with pd.ExcelWriter(path) as writer:
        pd.DataFrame(list(kpis.items()), columns=["KPI", "Value"]).to_excel(
            writer, sheet_name="Summary", index=False)
        by_product.to_excel(writer, sheet_name="By Product", index=False)
        by_region.to_excel(writer, sheet_name="By Region", index=False)
        by_seller.to_excel(writer, sheet_name="By Seller", index=False)
        by_month.to_excel(writer, sheet_name="Monthly", index=False)

        wb = writer.book
        _format_sheet(wb["Summary"])
        ws_summary = wb["Summary"]
        MONEY_KPIS = ("Total Revenue", "Average Order Value")
        COUNT_KPIS = ("Total Units", "Total Orders")
        for label_cell, value_cell in ws_summary.iter_rows(min_row=2, max_col=2):
            if label_cell.value in MONEY_KPIS:
                value_cell.number_format = CURRENCY_FMT
            elif label_cell.value in COUNT_KPIS:
                value_cell.number_format = "#,##0"
        for name in ("By Product", "By Region", "By Seller", "Monthly"):
            _format_sheet(wb[name], currency_cols=("B",))

        ws = wb["Monthly"]
        chart = BarChart()
        chart.title = "Monthly Revenue"
        chart.y_axis.title = "Revenue"
        chart.x_axis.title = "Month"
        data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width, chart.height = 24, 10
        ws.add_chart(chart, "F2")
    print(f"Report saved to {path}")

def main():
    df = load_clean_data()
    kpis = compute_kpis(df)
    by_product = aggregate_by(df, "Product")
    by_region = aggregate_by(df, "Region")
    by_month = monthly_trend(df)
    by_seller = aggregate_by(df, "Seller")

    for label, value in kpis.items():
        print(f"{label}: {value}")
    print(); print(by_product.head().to_string())

    build_report(df=df, kpis=kpis,
             by_product=by_product, by_region=by_region,
             by_seller=by_seller, by_month=by_month)


if __name__ == "__main__":
    main()